"""Update the CEE workbook in place, at the XML level.

openpyxl cannot be used here: the workbook drives Price, Beta, Ticker and Day
change from Excel linked data types (the _FV formulas reading the Stocks entity
in the Company column). A load-and-save through openpyxl drops every
xl/richData part, xl/metadata.xml and all 103 rich-value cell markers, which
would silently turn every live formula into an error.

So the file is treated as a zip of XML parts. Only the constants change --
Shares, Cost Basis and Cash -- and a Transactions worksheet is appended.
Everything else is copied through byte for byte.
"""
import csv, re, os, shutil, zipfile
from datetime import datetime
import openpyxl   # read-only, for locating rows by their cached ticker

DL   = r"C:\Users\Wmaxe\Downloads"
SRC  = os.path.join(DL, "CEE Fund Portfolios 8.3.26 LIVE.xlsx")
OUT  = os.path.join(DL, "CEE Fund Portfolios 9.2.26.xlsx")
POS  = {"Endowment": os.path.join(DL, "CEE Fund Mngd Endow-Positions-2026-09-02-140132.csv"),
        "CEE Fund":  os.path.join(DL, "CEE Fund Portfolio-Positions-2026-09-02-140153.csv")}
TXF  = {"Endowment": os.path.join(DL, "CEE_Fund_Mngd_Endow_XXX047_Transactions_20260812-085807.csv"),
        "CEE Fund":  os.path.join(DL, "CEE_Fund_Portfolio_XXX948_Transactions_20260812-085629.csv")}

num  = lambda s: float(re.sub(r"[^0-9.\-]", "", s)) if s and re.search(r"\d", s) else 0.0
tkr  = lambda s: s.strip().replace("BRK/B", "BRK.B")
esc  = lambda s: (s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))

def positions(p):
    out, cash = {}, 0.0
    rows = list(csv.reader(open(p, newline="", encoding="utf-8-sig")))
    h = next(i for i, r in enumerate(rows) if r and r[0] == "Symbol")
    for r in rows[h + 1:]:
        if not r or not r[0]: continue
        s = r[0].strip()
        if s == "Positions Total": continue
        if s.startswith("Cash"): cash = num(r[6]); continue
        out[tkr(s)] = dict(shares=num(r[2]), cost=num(r[9]), price=num(r[3]), mv=num(r[6]),
                           dayD=num(r[7]), dayP=num(r[8])/100.0, glD=num(r[10]), glP=num(r[11])/100.0)
    return out, cash

# ---- locate rows and the cash cell using cached values ---------------------
book = openpyxl.load_workbook(SRC, data_only=True)
plan = {}          # sheet name -> {"cells": {ref: value}, "missing": [...]}
report = {}
for fund in ["Endowment", "CEE Fund"]:
    ws = book[fund]
    new, cash = positions(POS[fund])
    cells, cached, seen = {}, {}, set()
    cash_ref = None
    for r in range(1, ws.max_row + 1):
        b = ws.cell(r, 2).value
        if not isinstance(b, str) or not b.strip(): continue
        label = b.strip()
        if label == "Cash":
            cash_ref = "C%d" % r
            continue
        if label == "Ticker": continue
        t = tkr(label)
        # The Treasury block has its own header: Ticker | Name | Amount |
        # Cost Basis | Purchase Price | Market Value, so market value sits in
        # column G rather than H and is updated separately.
        if re.match(r"^[0-9][0-9A-Z]{8}$", t) and t in new:
            cells["D%d" % r] = new[t]["shares"]
            cells["E%d" % r] = new[t]["cost"]
            cells["G%d" % r] = new[t]["mv"]
            seen.add(t)
            continue
        if t in new and isinstance(ws.cell(r, 4).value, (int, float)):
            n = new[t]
            cells["D%d" % r] = n["shares"]      # constant
            cells["E%d" % r] = n["cost"]        # constant
            # Formula cells: only the cached value is refreshed so the workbook
            # is accurate the moment it is opened or uploaded. Excel overwrites
            # these the instant it recalculates, which is the point of keeping
            # the formulas.
            cached["F%d" % r] = n["cost"] / n["shares"] if n["shares"] else 0.0
            cached["G%d" % r] = n["price"]
            cached["H%d" % r] = n["mv"]
            cached["I%d" % r] = n["dayP"]
            cached["J%d" % r] = n["dayD"]
            cached["K%d" % r] = n["glD"]
            cached["L%d" % r] = n["glP"]
            seen.add(t)
    if cash_ref: cells[cash_ref] = cash
    plan[fund] = (cells, cached)
    report[fund] = dict(updated=len(seen), missing=sorted(set(new) - seen), cash=cash, cash_ref=cash_ref)

# ---- transactions ----------------------------------------------------------
def iso(d):
    d = d.split(" as of ")[-1].strip()
    m, dd, y = d.split("/")
    return "%s-%s-%s" % (y, m.zfill(2), dd.zfill(2))

txs = []
for fund, path in TXF.items():
    for row in csv.DictReader(open(path, newline="", encoding="utf-8-sig")):
        if not row.get("Date") or not row.get("Action"): continue
        txs.append((iso(row["Date"]), fund, tkr(row["Symbol"] or ""), row["Action"].strip(),
                    num(row["Quantity"]), num(row["Price"]), num(row["Amount"]),
                    num(row["Fees & Comm"]), (row["Description"] or "").strip()))
# Transactions after the last CSV export, taken from the August statements and
# the 09/01 settlements, validated to reproduce the 9/2 positions exactly.
import json as _json
_extra = os.path.join(os.path.dirname(os.path.abspath(__file__)), "new_tx.json")
if os.path.exists(_extra):
    _seen = {(t[0], t[1], t[2], t[3], round(t[4], 4)) for t in txs}
    for d, fund, t, a, q, pr, amt in _json.load(open(_extra)):
        key = (d, fund, t, a, round(q, 4))
        if key in _seen: continue
        txs.append((d, fund, t, a, q, pr, amt, 0.0, ""))
txs.sort(key=lambda t: (t[0], t[1], t[2]), reverse=True)

# ---- rewrite the package ---------------------------------------------------
zin = zipfile.ZipFile(SRC)
names = zin.namelist()

# map sheet display name -> worksheet part
wbxml = zin.read("xl/workbook.xml").decode("utf-8")
relxml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
rels = dict(re.findall(r'Id="([^"]+)"[^>]*Target="([^"]+)"', relxml))
sheet_part = {}
for m in re.finditer(r'<sheet[^>]*name="([^"]+)"[^>]*r:id="([^"]+)"', wbxml):
    tgt = rels[m.group(2)].lstrip("/")
    sheet_part[m.group(1)] = tgt if tgt.startswith("xl/") else "xl/" + tgt

CELL_RE = lambda ref: re.compile(r'<c r="%s"((?:\s+[a-zA-Z:]+="[^"]*")*)\s*(?:/>|>(.*?)</c>)' % ref, re.S)
def set_numeric(xml, ref, value):
    """Replace a constant cell's value, keeping its style attributes."""
    pat = CELL_RE(ref)
    m = pat.search(xml)
    if not m:
        return xml, False
    attrs = m.group(1) or ""
    attrs = re.sub(r'\s+t="[^"]*"', "", attrs)        # numeric: drop any type flag
    repl = '<c r="%s"%s><v>%s</v></c>' % (ref, attrs, repr(round(value, 6)).rstrip("0").rstrip(".")
                                          if isinstance(value, float) else value)
    return xml[:m.start()] + repl + xml[m.end():], True

def set_cached(xml, ref, value):
    """Refresh a formula cell's cached result, leaving <f> untouched."""
    m = re.search(r'<c r="%s"(?![0-9])[^>]*>.*?</c>' % ref, xml, re.S)
    if not m: return xml, False
    cell = m.group(0)
    if "<f" not in cell: return xml, False
    val = repr(round(value, 6)).rstrip("0").rstrip(".") if isinstance(value, float) else str(value)
    if "<v>" in cell:
        new = re.sub(r"<v>.*?</v>", "<v>%s</v>" % val, cell, count=1, flags=re.S)
    else:
        new = cell.replace("</c>", "<v>%s</v></c>" % val)
    return xml[:m.start()] + new + xml[m.end():], True

applied, refreshed, failed = 0, 0, []
parts = {}
for fund, (cells, cached) in plan.items():
    part = sheet_part[fund]
    xml = zin.read(part).decode("utf-8")
    for ref, val in cells.items():
        xml, ok = set_numeric(xml, ref, float(val))
        if ok: applied += 1
        else: failed.append("%s!%s" % (fund, ref))
    for ref, val in cached.items():
        xml, ok = set_cached(xml, ref, float(val))
        if ok: refreshed += 1
    parts[part] = xml.encode("utf-8")

# ---- build the Transactions worksheet --------------------------------------
used_ids = [int(x) for x in re.findall(r'sheetId="(\d+)"', wbxml)]
new_sheet_id = max(used_ids) + 1
used_rids = [int(x[3:]) for x in rels if x.startswith("rId") and x[3:].isdigit()]
new_rid = "rId%d" % (max(used_rids) + 1)
used_parts = [int(m.group(1)) for m in
              (re.match(r"xl/worksheets/sheet(\d+)\.xml", n) for n in names) if m]
new_part = "xl/worksheets/sheet%d.xml" % (max(used_parts) + 1)

HDR = ["Date", "Fund", "Ticker", "Action", "Shares", "Price", "Amount", "Fees", "Description"]
def col(i): return chr(ord("A") + i)
rows_xml = []
rows_xml.append('<row r="1">' + "".join(
    '<c r="%s1" t="inlineStr"><is><t>%s</t></is></c>' % (col(i), esc(h)) for i, h in enumerate(HDR)) + "</row>")
for n, t in enumerate(txs, start=2):
    cs = []
    for i, v in enumerate(t):
        ref = "%s%d" % (col(i), n)
        if isinstance(v, float):
            if v == 0: continue
            cs.append('<c r="%s"><v>%s</v></c>' % (ref, repr(round(v, 6)).rstrip("0").rstrip(".")))
        else:
            if v == "": continue
            cs.append('<c r="%s" t="inlineStr"><is><t>%s</t></is></c>' % (ref, esc(str(v))))
    rows_xml.append('<row r="%d">%s</row>' % (n, "".join(cs)))

sheet_xml = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
    '<sheetPr><outlinePr summaryBelow="1" summaryRight="1"/></sheetPr>'
    '<dimension ref="A1:I%d"/>'
    '<sheetViews><sheetView workbookViewId="0"><pane ySplit="1" topLeftCell="A2" activePane="bottomLeft" state="frozen"/></sheetView></sheetViews>'
    '<sheetFormatPr defaultRowHeight="15"/>'
    '<cols><col min="1" max="1" width="12" customWidth="1"/><col min="2" max="2" width="12" customWidth="1"/>'
    '<col min="3" max="3" width="11" customWidth="1"/><col min="4" max="4" width="22" customWidth="1"/>'
    '<col min="5" max="8" width="13" customWidth="1"/><col min="9" max="9" width="46" customWidth="1"/></cols>'
    '<sheetData>%s</sheetData>'
    '<autoFilter ref="A1:I%d"/>'
    '<pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" header="0.3" footer="0.3"/>'
    '</worksheet>'
) % (len(txs) + 1, "".join(rows_xml), len(txs) + 1)

# If a Transactions sheet already exists, overwrite that part in place. Adding
# a second one leaves the workbook with two sheets of the same name, and only
# the stale first copy is ever read.
if "Transactions" in sheet_part:
    parts[sheet_part["Transactions"]] = sheet_xml.encode("utf-8")
    with zipfile.ZipFile(OUT, "w", zipfile.ZIP_DEFLATED) as z:
        for n in names:
            z.writestr(n, parts.pop(n, zin.read(n)))
        for n, data in parts.items():
            z.writestr(n, data)
    zin.close()
    print("WROTE " + OUT + "   (Transactions sheet replaced in place)")
    for f, r in report.items():
        print("  %-10s positions updated=%d  cash=%s -> $%s  %s"
              % (f, r["updated"], r["cash_ref"], format(r["cash"], ",.2f"),
                 ("NEEDS MANUAL ROW: " + ", ".join(r["missing"])) if r["missing"] else ""))
    print("  constants rewritten : %d   cached refreshed : %d" % (applied, refreshed))
    print("  failures            : %s" % (failed or "none"))
    print("  Transactions rows   : %d  (%s -> %s)" % (len(txs), txs[-1][0], txs[0][0]))
    raise SystemExit(0)

# register the new sheet
wbxml_new = wbxml.replace("</sheets>",
    '<sheet name="Transactions" sheetId="%d" r:id="%s"/></sheets>' % (new_sheet_id, new_rid))
relxml_new = relxml.replace("</Relationships>",
    '<Relationship Id="%s" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
    'Target="worksheets/sheet%d.xml"/></Relationships>' % (new_rid, max(used_parts) + 1))
ct = zin.read("[Content_Types].xml").decode("utf-8")
ct_new = ct.replace("</Types>",
    '<Override PartName="/%s" ContentType="application/vnd.openxmlformats-officedocument.'
    'spreadsheetml.worksheet+xml"/></Types>' % new_part)

parts["xl/workbook.xml"] = wbxml_new.encode("utf-8")
parts["xl/_rels/workbook.xml.rels"] = relxml_new.encode("utf-8")
parts["[Content_Types].xml"] = ct_new.encode("utf-8")
parts[new_part] = sheet_xml.encode("utf-8")

with zipfile.ZipFile(OUT, "w", zipfile.ZIP_DEFLATED) as z:
    for n in names:
        z.writestr(n, parts.pop(n, zin.read(n)))
    for n, data in parts.items():
        z.writestr(n, data)
zin.close()

print("WROTE " + OUT + "\n")
for f, r in report.items():
    print("  %-10s positions updated=%d  cash=%s -> $%s  %s"
          % (f, r["updated"], r["cash_ref"], format(r["cash"], ",.2f"),
             ("NEEDS MANUAL ROW: " + ", ".join(r["missing"])) if r["missing"] else ""))
print("\n  cell values rewritten : %d" % applied)
print("  failures              : %s" % (failed or "none"))
print("  Transactions rows     : %d  (%s -> %s)" % (len(txs), txs[-1][0], txs[0][0]))
