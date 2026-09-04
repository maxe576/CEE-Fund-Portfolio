# Rebuild the CEE workbook from the 2026-08-03 Schwab position files, keeping the
# exact sheet layout the dashboard parser expects, and add a Transactions tab
# holding the full Schwab history so the workbook becomes the single master file.
import csv, re, os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

DL = r"C:\Users\Wmaxe\Downloads"
OLD = os.path.join(DL, "CEE Fund Portfolios 4.12.26 (5).xlsx")
POS = {"Endowment": os.path.join(DL, "CEE Fund Mngd Endow-Positions-2026-08-03-113757.csv"),
       "CEE Fund":  os.path.join(DL, "CEE Fund Portfolio-Positions-2026-08-03-113807.csv")}
TXF = {"Endowment": os.path.join(DL, "CEE_Fund_Mngd_Endow_XXX047_Transactions_20260812-085807.csv"),
       "CEE Fund":  os.path.join(DL, "CEE_Fund_Portfolio_XXX948_Transactions_20260812-085629.csv")}

def num(s):
    if not s or not re.search(r"\d", s): return 0.0
    return float(re.sub(r"[^0-9.\-]", "", s))
def pct(s):
    return num(s) / 100.0 if s and "%" in s else 0.0
def tkr(s):
    return s.strip().replace("BRK/B", "BRK.B")

# ---- carry sector + beta forward from the existing workbook ----------------
old = openpyxl.load_workbook(OLD, data_only=True)
meta = {}
for sh in old.sheetnames:
    ws = old[sh]
    for r in range(1, ws.max_row + 1):
        t = ws.cell(r, 2).value
        if not isinstance(t, str) or not t.strip(): continue
        sec, beta = ws.cell(r, 13).value, ws.cell(r, 14).value
        if isinstance(sec, str) and sec.strip() and isinstance(beta, (int, float)):
            meta[tkr(t)] = (sec.strip(), float(beta))

# MBGL is new: an ExxonMobil spin-off received 07/01/2026. Parent sector is used
# as the default and flagged below so the committee can set it deliberately.
NEW_META = {"MBGL": ("Energy", 1.0)}
FLAGGED = []

def read_positions(path):
    eq, etf, fixed, cash = [], [], [], 0.0
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    hdr = next(i for i, r in enumerate(rows) if r and r[0] == "Symbol")
    for r in rows[hdr + 1:]:
        if not r or not r[0]: continue
        sym = r[0].strip()
        if sym == "Positions Total": continue
        if sym.startswith("Cash"):
            cash = num(r[6]); continue
        rec = dict(ticker=tkr(sym), desc=r[1].strip(), shares=num(r[2]), price=num(r[3]),
                   mv=num(r[6]), dayD=num(r[7]), dayP=pct(r[8]), cost=num(r[9]),
                   glD=num(r[10]), glP=pct(r[11]), atype=r[14].strip())
        rec["avg"] = rec["cost"] / rec["shares"] if rec["shares"] else 0.0
        if rec["atype"].startswith("Fixed"): fixed.append(rec)
        elif rec["atype"].startswith("ETF"): etf.append(rec)
        else: eq.append(rec)
    return eq, etf, fixed, cash

def sector_beta(t, is_etf):
    if t in meta: return meta[t]
    if t in NEW_META:
        FLAGGED.append(t); return NEW_META[t]
    FLAGGED.append(t)
    return ("ETF" if is_etf else "Unknown", 1.0)

HDR_FILL = PatternFill("solid", fgColor="0C2A5E")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE    = Font(bold=True, size=13, color="0C2A5E")
SECTION  = Font(bold=True, size=11, color="0C2A5E")
MONEY, PCT4, SHR = "#,##0.00", "0.00%", "#,##0.0000"

wb = openpyxl.Workbook(); wb.remove(wb.active)
COLS = ["Ticker","Company","Shares","Cost Basis","Avg Price","Market Price","Market Value",
        "Day %","Day $","G/L $","G/L %","Sector","Beta"]
summary = {}

for fund in ["Endowment", "CEE Fund"]:
    eq, etf, fixed, cash = read_positions(POS[fund])
    eq.sort(key=lambda x: (sector_beta(x["ticker"], False)[0], x["ticker"]))
    etf.sort(key=lambda x: x["ticker"])
    ws = wb.create_sheet(fund)
    # The dashboard reads the ticker from column B. sheet_to_json indexes from the
    # sheet used range, so column A must carry content or every column shifts left
    # by one and company names get read as tickers.
    ws.cell(1, 1, "Captains Educational Enrichment Fund").font = TITLE
    ws.cell(2, 1, "Positions as of 2026-08-03  -  " + fund).font = Font(italic=True, size=9, color="666666")
    ws.cell(5, 1, " ")   # anchors the used range at column A

    def header_block(row, label, etf_names=False):
        ws.cell(row, 2, label).font = SECTION
        hr = row + 2
        for i, h in enumerate(COLS):
            txt = "ETF Name" if (i == 1 and etf_names) else h
            c = ws.cell(hr, 2 + i, txt)
            c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = Alignment(horizontal="center")
        return hr + 1

    def write_rows(start, items, is_etf):
        r = start
        for it in items:
            sec, beta = sector_beta(it["ticker"], is_etf)
            vals = [it["ticker"], it["desc"], it["shares"], it["cost"], it["avg"], it["price"],
                    it["mv"], it["dayP"], it["dayD"], it["glD"], it["glP"], sec, beta]
            for i, v in enumerate(vals):
                c = ws.cell(r, 2 + i, v)
                if i == 2: c.number_format = SHR
                elif i in (3, 4, 5, 6, 8, 9): c.number_format = MONEY
                elif i in (7, 10): c.number_format = PCT4
                elif i == 12: c.number_format = "0.0000"
            r += 1
        return r

    r = header_block(3, fund + " Equities")
    r = write_rows(r, eq, False)
    r = header_block(r + 1, fund + " ETF" + chr(39) + "s", etf_names=True)
    r = write_rows(r, etf, True)
    if fixed:
        ws.cell(r + 1, 2, fund + " Treasury Stock").font = SECTION
        hr = r + 3
        for i, h in enumerate(["Ticker","Name","Amount","Cost Basis","Avg Price","Market Price","Market Value"]):
            c = ws.cell(hr, 2 + i, h); c.font = HDR_FONT; c.fill = HDR_FILL
        r = hr + 1
        for it in fixed:
            for i, v in enumerate([it["ticker"], it["desc"], it["shares"], it["cost"],
                                   it["avg"], it["price"], it["mv"]]):
                c = ws.cell(r, 2 + i, v)
                if i >= 3: c.number_format = MONEY
            r += 1
    ws.cell(r + 1, 2, "Cash").font = Font(bold=True)
    ws.cell(r + 1, 3, cash).number_format = MONEY
    ws.column_dimensions["A"].width = 3
    for col, w in zip("BCDEFGHIJKLMN", [11,34,13,13,13,13,15,10,11,13,10,24,9]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "B6"
    summary[fund] = dict(eq=len(eq), etf=len(etf), fixed=len(fixed), cash=cash,
                         mv=sum(x["mv"] for x in eq + etf + fixed))

# ---- Transactions tab ------------------------------------------------------
ws = wb.create_sheet("Transactions")
THDR = ["Date","Fund","Ticker","Action","Shares","Price","Amount","Fees","Description"]
note = ("Full transaction history, both accounts. Columns A-F drive the dashboard upload; "
        "Amount carries dividends, interest and wires that have no share count.")
ws.cell(1, 1, note).font = Font(italic=True, size=9, color="666666")
for i, h in enumerate(THDR):
    c = ws.cell(2, 1 + i, h); c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = Alignment(horizontal="center")

def iso(d):
    d = d.split(" as of ")[-1].strip()
    m, dd, y = d.split("/")
    return datetime(int(y), int(m), int(dd))

txs = []
for fund, path in TXF.items():
    with open(path, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            if not row.get("Date") or not row.get("Action"): continue
            txs.append(dict(d=iso(row["Date"]), fund=fund, t=tkr(row["Symbol"] or ""),
                            a=row["Action"].strip(), q=num(row["Quantity"]), p=num(row["Price"]),
                            m=num(row["Amount"]), fee=num(row["Fees & Comm"]),
                            desc=(row["Description"] or "").strip()))
txs.sort(key=lambda x: (x["d"], x["fund"], x["t"]), reverse=True)
for i, t in enumerate(txs):
    r = 3 + i
    ws.cell(r, 1, t["d"]).number_format = "yyyy-mm-dd"
    ws.cell(r, 2, t["fund"]); ws.cell(r, 3, t["t"] or None); ws.cell(r, 4, t["a"])
    if t["q"]: ws.cell(r, 5, t["q"]).number_format = SHR
    if t["p"]: ws.cell(r, 6, t["p"]).number_format = MONEY
    if t["m"]: ws.cell(r, 7, t["m"]).number_format = MONEY
    if t["fee"]: ws.cell(r, 8, t["fee"]).number_format = MONEY
    ws.cell(r, 9, t["desc"])
for col, w in zip("ABCDEFGHI", [12,12,11,22,13,12,13,9,46]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A3"
ws.auto_filter.ref = "A2:I" + str(2 + len(txs))

OUT = os.path.join(DL, "CEE Fund Portfolios 8.3.26.xlsx")
wb.save(OUT)
print("WROTE " + OUT + "\n")
for f, s in summary.items():
    print("  %-10s equities=%3d etfs=%3d fixed=%d cash=$%10s securities=$%12s total=$%12s"
          % (f, s["eq"], s["etf"], s["fixed"], format(s["cash"], ",.2f"),
             format(s["mv"], ",.2f"), format(s["mv"] + s["cash"], ",.2f")))
print("\n  Transactions rows: %d  (%s -> %s)" % (len(txs), txs[-1]["d"].date(), txs[0]["d"].date()))
print("  Needing a sector/beta decision: %s" % (sorted(set(FLAGGED)) or "none"))
